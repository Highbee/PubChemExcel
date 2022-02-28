import pubchempy as pcp
import openpyxl as op

wb = op.load_workbook('compounds_library_use.xlsx')
sheet = wb["Sheet1"]
max_row = sheet.max_row
count = 0
compound_not_found = []

for rw in range(1, max_row+1):
    compound_name = sheet.cell(row=rw, column=1).value
    compound_cid = sheet.cell(row=rw, column=2).value
    try:
        pcp.download(
            "SDF", f"downloads\{compound_name}_@_{compound_cid}.sdf", compound_cid, 'cid', record_type='3d')
        print(f"{compound_name} downloaded")
        count += 1
    except (pcp.NotFoundError):
        message = f"{compound_name}, with cid: {compound_cid} "
        compound_not_found.append(message)
print(f"Successfully downloaded {count} files in sdf format")
if compound_not_found:
    print("The following compounds were not available in the format specified or do not exist")
    for c in compound_not_found:
        print(c)
